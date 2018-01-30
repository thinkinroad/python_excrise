from openpyxl import *
def get_url():
    url_=[]
    workbook_ = load_workbook(filename='1.xlsx')
    sheetnames =workbook_.get_sheet_names()
    sheet = workbook_.get_sheet_by_name(sheetnames[0])

    for rowNum in range(1,24):
        url_.append(sheet.cell(row=rowNum,column=3).value)
    return url_
if __name__ == '__main__':
    url_ = get_url()
    print(url_)